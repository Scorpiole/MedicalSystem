# 定义一个管理系统类
import matplotlib.image as im
import matplotlib.pyplot as plt
import shutil
import os
import openpyxl
import re
from pathlib import *

patience_list = []  # 创建保存病人ID所用的列表
patience_dict = {}

class PatienceManager(object):
    def __init__(self, path1='', path2='', path3='F:\桌面软件\学习\课设4\课程设计检验数据.xlsx'):
        self.Local_Path = path1
        self.Target_Path = path2
        self.Excel_Path = path3
        self.sheet_name = ''
        self.init_patience_list()
    # 将所有患者的ID写入列表,初始化ID列表
    def init_patience_list(self):
        patience_dict.clear()
        patience_list.clear()
        book = openpyxl.load_workbook(self.Excel_Path)  # 创建excel操作对象
        sheet_list = book.sheetnames
        self.sheet_name = sheet_list[0]
        sheet = book[self.sheet_name]
        rows = sheet.max_row
        self.cols = sheet.max_column
        for i in range(1, rows-1):
            cell_value = str(sheet.cell(row=i+2, column=1).value)
            patience_dict[i] = cell_value
            patience_list.append(cell_value)
        book.close()
        print(patience_dict)

    # 程序入口函数
    def run(self):
        while True:    # 进行遍历
            # 1 显示功能菜单
            self.show_menu()  # 调用对应函数

            # 2 接受用户输入
            menu_num = int(input('请输入您需要的功能序列号：'))

            # 3 根据用户输入的序列进行对应功能的执行
            if menu_num == 1:
                # 增加病患信息
                self.add_patience()
            elif menu_num == 2:
                # 删除病患信息
                self.del_patience()
            elif menu_num == 3:
                # 修改病患信息
                self.modify_patience()
            elif menu_num == 4:
                # 查询病患信息
                self.search_patience()
            elif menu_num == 5:
                # 显示所有病患信息
                self.show_patience()
            elif menu_num == 6:
                # ID查找病患图像
                self.search_picture()
            elif menu_num == 7:
                # 保存病患图像至本地
                self.save_picture()
            elif menu_num == 8:
                # 退出医疗信息系统
                break

    # 完成入口函数内各函数的功能
    @staticmethod
    # 显示功能菜单
    def show_menu():
        print('------------------------')
        print('请选择您所需要的功能：')
        print('1.增加病患信息')
        print('2.删除病患信息')
        print('3.修改病患信息')
        print('4.查询病患信息')
        print('5.显示所有入院病患信息')
        print('6.查找病患图像')
        print('7.保存病患图像')
        print('8.使用结束，退出系统')
        print('------------------------')

    # 增加病患信息
    def add_patience(self, list1):
        book = openpyxl.load_workbook(self.Excel_Path)  # 创建excel操作对象
        sheet = book[self.sheet_name]
        rows = sheet.max_row  # 获取目前excel表格的总行数
        print('添加的数据所在的行是：', rows+1, '病患信息是：', list1)
        sheet.cell(rows + 1, 1).value = list1[0]
        sheet.cell(rows + 1, 2).value = list1[1]
        sheet.cell(rows + 1, 3).value = list1[2]
        sheet.cell(rows + 1, 4).value = list1[3]
        sheet.cell(rows + 1, 5).value = list1[4]
        sheet.cell(rows + 1, 6).value = list1[5]
        sheet.cell(rows + 1, 7).value = list1[6]
        sheet.cell(rows + 1, 8).value = list1[7]
        sheet.cell(rows + 1, 9).value = list1[8]
        sheet.cell(rows + 1, 10).value = list1[9]
        sheet.cell(rows + 1, 11).value = list1[10]
        sheet.cell(rows + 1, 12).value = list1[11]
        sheet.cell(rows + 1, 13).value = list1[12]
        sheet.cell(rows + 1, 14).value = list1[13]
        sheet.cell(rows + 1, 15).value = list1[14]
        sheet.cell(rows + 1, 16).value = list1[15]
        sheet.cell(rows + 1, 17).value = list1[16]
        sheet.cell(rows + 1, 18).value = list1[17]
        sheet.cell(rows + 1, 19).value = list1[18]
        sheet.cell(rows + 1, 20).value = list1[19]
        sheet.cell(rows + 1, 21).value = list1[20]
        sheet.cell(rows + 1, 22).value = list1[21]
        sheet.cell(rows + 1, 23).value = list1[22]
        sheet.cell(rows + 1, 24).value = list1[23]
        book.save(self.Excel_Path)
        cell_value = str(sheet.cell(row=rows+1, column=1).value)
        patience_list.append(cell_value)  # 更新ID列表
        patience_dict[rows-1] = cell_value  # 更新ID字典
        book.close()

    # 删除病患信息
    def del_patience(self, ID):
        book = openpyxl.load_workbook(self.Excel_Path)  # 创建excel操作对象
        sheet = book[self.sheet_name]
        # ID=str(id)
        patience_ID_row = patience_list.index(ID)+3  # 列表方式，获取病患ID的行值
        # patience_ID_row = 0
        # for key, value in patience_dict.items():  # 字典方式，获取病患ID的行值
        #     if value == ID:
        #         patience_ID_row = key + 2
        print('删除的行是：', patience_ID_row - 2)
        sheet.delete_rows(patience_ID_row)  # 通过ID列表删除对应ID所在行
        # sheet.delete_rows(patience_ID_row)  # 通过ID字典删除对应ID所在行
        book.save(self.Excel_Path)
        patience_list.remove(ID)  # 更新ID列表
        book.close()

    # 修改病患信息
    def modify_patience(self, list1):
        book = openpyxl.load_workbook(self.Excel_Path)  # 创建excel操作对象
        sheet = book[self.sheet_name]
        rows = sheet.max_row  # 获取目前excel表格的总行数
        # modify_list = [id,num1,num2,num3,num4,num5,num6,num7,num8,num9,num10,num11,num12,num13,num14,num15,num16,num17,num18,num19,num20,num21,num22]
        modify_index = []  # 存放要修改的变量的下标
        for i in range(1, 24):
            if list1[i] != 'None':
               modify_index.append(str(i + 1))  # 存入第几列需要修改
        patience_ID_row = patience_list.index(list1[0]) + 1  # 获取病患ID的行值
        count = len(modify_index)  # 记录需要修改的元素个数
        for j in range(count):
            sheet.cell(patience_ID_row + 2, int(modify_index[j])).value = list1[int(modify_index[j]) - 1]  # 进行修改
        book.save(self.Excel_Path)
        book.close()

    # 查询病患信息
    def search_patience(self, ID):
        lista = []
        book = openpyxl.load_workbook(self.Excel_Path)  # 创建excel操作对象
        sheet = book[self.sheet_name]
        # ID = str(input('请输入病患ID：'))
        # ID=str(id)
        patience_ID_row = patience_list.index(ID) + 3  # 获取病患ID的行值
        for i in range(1, sheet.max_column+1):
            if i == 2:
                time = str(sheet.cell(patience_ID_row, i).value)
                time = time[0:10]
                lista.append(time)
            else:
                lista.append(str(sheet.cell(patience_ID_row, i).value))
            # print(str(sheet.cell(patience_ID_row, i).value)+' ', end='')
        book.close()
        # print()
        return lista

    # 显示所有入院病患信息
    def show_patience(self):
        book = openpyxl.load_workbook(self.Excel_Path)  # 创建excel操作对象
        sheet = book[self.sheet_name]
        for i in range(1, sheet.max_row+1):
            for j in range(1, sheet.max_column+1):
               print(str(sheet.cell(i, j).value) + '       ', end='')
            print()
        book.close()

    # ID查找病患图像
    def search_picture(self, pid, sIs_ill):
        if sIs_ill != '0' and sIs_ill != '1':
            path_name = self.Local_Path + '/no_design'
        else:
            path_name = self.Local_Path + '/M' + sIs_ill
        item_ID = str(pid)
        img_dir = Path(path_name)
        if img_dir.is_dir():
            for item in os.listdir(path=path_name):
                if item_ID == item.split('-')[0]:
                    return 1
        return 0

    # ID删除病患图像
    def del_picture(self, pid, sIsill):
        sign = 0  # 标记
        if sIsill != '0' and sIsill != '1':
            path_name = self.Local_Path + '/no_design'  # 加载本地路径读入图片
        else:
            path_name = self.Local_Path + '/M' + sIsill
        item_ID = str(pid)
        img_dir = Path(path_name)
        if img_dir.is_dir():
            for item in os.listdir(path=path_name):
                if item_ID == item.split('-')[0]:
                    os.remove(os.path.join(path_name, item))
                    sign = 1  # 标记
                    break
        if sign == 0:
            print('您输入的ID暂无图像数据。')

    # 保存病患图像至本地
    def save_picture(self, new_img_path, new_id, isill):
        # root_id 图片在原文件夹中的名称
        # new_id 图片复制到目标文件夹的名称
        listA = new_img_path.split("/")
        root_id = listA[-1]
        root_path = listA[0]

        for i in range(1, len(listA) - 1):
            root_path = root_path + '/' + listA[i]
        print(isill)
        print('录入新图片！')
        # root_path = self.Target_Path  # 从该文件夹读取图片
        if isill != '0' and isill != '1':
            save_path = self.Local_Path + '/no_design'
            if not os.path.exists(save_path):
                os.mkdir(os.path.join(self.Local_Path, 'no_design'))
        else:
            save_path = self.Local_Path + '/M'+str(isill)   # 将外部文件夹图片复制到该本地文件夹保存
        print(save_path)
        if str(root_id) in os.listdir(path=root_path):
            if str(new_id) + '.jpg' in os.listdir(path=save_path):
                print('本地文件夹中已存在该ID图像。')
            else:
                shutil.copy(os.path.join(root_path, str(root_id)), os.path.join(save_path, str(new_id)+'.jpg'))
        else:
            print('原文件夹中无该图像，请输入正确名称。')

    # 判断字符是否是数字
    def is_number(self,s):
        try:
            float(s)
            return True
        except ValueError:
            pass
        try:
            import unicodedata
            unicodedata.numeric(s)
            return True
        except (TypeError, ValueError):
            pass
        return False

    # 检测数据是否输入正常
    def check_data(self, data_list):
        check_sign = [0] * 22  # 初始化标记列表为全0列表，数据合理则录入0，不合理录入1
        name = ['入院日期', '性别', '年龄', '身高', 'BMI', '糖尿病史', '高血压史', '肾病家族史', '是否服用ACEI/ARB', '肾穿前尿蛋白史', '持续镜下血尿史', '收缩压',
                '舒张压', '右肾长度', '右肾宽度', '右肾高度', '右肾皮质厚度', '左肾长度', '左肾宽度', '左肾高度', '左肾皮质厚度', '病例分级']
        # 日期
        big_month = [1, 3, 5, 7, 8, 10, 12]  # 大月有31日
        print('输入的日期是', data_list[0])
        if data_list[0] == 'None':
            check_sign[0] = 0
        elif '-' in data_list[0] or '年' in data_list[0] or '.' in data_list[0]:
            if '-' in data_list[0]:
                date_split_list = data_list[0].split('-')
            elif '.' in data_list[0]:
                date_split_list = data_list[0].split('.')
                print('111')
            else:
                date_split_list = re.split("[年 月 日]", data_list[0])
                date_split_list.pop()
                print('分割后列表', date_split_list)
            if len(date_split_list) == 3:
                year = date_split_list[0]
                month = date_split_list[1]
                day = date_split_list[2]
                # 闰年：能被4整除但不能被100整除 或 能被400整除，满足以上两条件中任一条件的即为闰年，此时2月有29天
                if month != '2':  # 非2月
                    check_sign[0] = 0 if (
                            year.isdigit() and (int(year) <= 2022) and (int(year) >= 2000) and month.isdigit() and
                            (1 <= int(month) <= 12) and day.isdigit() and (1 <= int(day) <= (31 if (int(month) in big_month)
                                                                                             else 30))) else 1
                else:  # 2月
                    if (int(year) % 4 == 0 and int(year) % 100 != 0) or (int(year) % 400 == 0):  # 闰年2月，day<=29
                        check_sign[0] = 0 if (
                                year.isdigit() and (int(year) <= 2022) and (int(year) >= 2000) and day.isdigit() and (
                                        int(month) >= 1) and (int(day) <= 29) and (int(day) >= 1)) else 1
                        print('闰年')
                    else:  # 非闰年2月，day<=28
                        check_sign[0] = 0 if (
                                year.isdigit() and (int(year) <= 2022) and (int(year) >= 2000) and day.isdigit() and (
                                        int(month) >= 1) and (int(day) <= 28) and (int(day) >= 1)) else 1
                        print('平年')
            else:
                check_sign[0] = 1
        else:
            check_sign[0] = 1
        # 性别
        check_sign[1] = 0 if ((data_list[1] == '男') or (data_list[1] == '女') or (data_list[1] == 'None')
                              or (data_list[1] == '')) else 1
        # 糖尿病史，高血压史，肾病家族史，是否服用ACEI/ARB均输入0/1即可，在check_sign中对应check_sign[5]-check_sign[8]
        for i in range(5, 9):
            check_sign[i] = 0 if ((data_list[i] == '0') or (data_list[i] == '1') or (data_list[i] == 'None')
                                  or (data_list[i] == '')) else 1
        # 年龄,身高,BMI,肾穿前尿蛋白史(月数),持续镜下血尿史(月数),收缩压,舒张压,右肾长度,右肾宽度,右肾高度,右肾皮质厚度,左肾长度,左肾宽度,左肾高度,左肾皮质厚度共15个变量应设置不同的合理值区间
        index = [2, 3, 4, 9, 10, 11, 12, 13, 14, 15, 16, 17, 18, 19, 20]  # 以上15个变量对应于check_sign的下标
        min_vaule = [0, 50, 10, 0, 0, 50, 50, 5, 3, 3, 0, 5, 3, 3, 0]  # 以上15个变量的合理值的最小值
        max_vaule = [110, 230, 35, 300, 300, 250, 250, 20, 10, 10, 4, 20, 10, 10, 4]  # 以上15个变量的合理值的最大值
        for j in range(0, 15):
            check_sign[index[j]] = 0 if (self.is_number(data_list[index[j]]) and (float(data_list[index[j]]) <= max_vaule[j]) and (
                        float(data_list[index[j]]) >= min_vaule[j]) or data_list[index[j]] == ''
                                         or data_list[index[j]] == 'None') else 1

        # 病例分级
        check_sign[21] = 0 if (data_list[21] == '0' or data_list[21] == '1' or data_list[21] == ''
                               or data_list[21] == 'None') else 1
        # 进行错误提示
        flase_name = []  # 放置错误变量
        for k in range(0, 22):
            if check_sign[k] == 1:
                flase_name.append(name[k])  # 检测到错误变量，放入flase_name中
        print('输入有误的相关项目：', flase_name)
        if flase_name:
            result = flase_name[0]
            for i in range(1, len(flase_name)):
                result = result + '、' + flase_name[i]
            print('您输入的' + str(flase_name) + '可能存在错误，请仔细确认后再次输入。')
            result_str = '您输入的' + result + '可能存在错误，请仔细确认后再次输入。'
            return '您输入的  ' + result + '  可能存在错误，请仔细确认后再次输入。'
        else:
            return 0

    def turn_date_2_chinese(self, date):
        chinese_date = ''


        return chinese_date




#-------------------9-----------------------------------------------------------------程序测试------------------------------------------------------
if __name__ == '__main__':
    test = PatienceManager()
    test.run()
